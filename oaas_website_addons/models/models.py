# -*- coding: utf-8 -*-
import io
import json

import openpyxl
from PIL import Image as PILImage

# openpyxl charge les images embarquees via PIL des load_workbook(). Certaines
# images du fichier Excel sont tres grandes (> 89 Mpx) et declenchent la
# protection "decompression bomb" de Pillow (warning, voire erreur selon la
# version). On RELEVE la limite plutot que de la desactiver (None) : on garde
# ainsi un garde-fou contre les vraies bombes (images de plusieurs Go), tout en
# laissant passer nos images legitimes. Les images sont ensuite redimensionnees
# a l'import (cf. add_attachment_file dans scripts/test_import_xls.py).
# 250 Mpx ~= 750 Mo en RAM decompresse : large pour nos cas, fini pour bloquer
# un fichier malveillant.
PILImage.MAX_IMAGE_PIXELS = 250_000_000
from odoo import models, fields, api,  _
from odoo.exceptions import UserError
import base64
import csv
from ..scripts.test_import_xls import convert_row_to_blog_post
from translate import Translator

# class oaas_website_addons(models.Model):
#     _name = 'oaas_website_addons.oaas_website_addons'
#     _description = 'oaas_website_addons.oaas_website_addons'

#     name = fields.Char()
#     value = fields.Integer()
#     value2 = fields.Float(compute="_value_pc", store=True)
#     description = fields.Text()
#
#     @api.depends('value')
#     def _value_pc(self):
#         for record in self:
#             record.value2 = float(record.value) / 100

class BlogPostImport(models.TransientModel):
    _name = 'blog.post.import'
    _description = 'Blog Post Import'

    file = fields.Binary(string=_('File'), required=True)
    filename = fields.Char(string=_('Filename'))
    description = fields.Char(string=_('Description de votre import'))
    when = fields.Datetime(string=_("Date d'enregistrement"), default=fields.Datetime.now())

    def import_data(self):
        if not self.file:
            raise UserError(_('Veuillez cliquez sur importer un fichier avant'))
        try:
            data = base64.b64decode(self.file)
            excel_buffer = io.BytesIO(data)
            pxl_doc = openpyxl.load_workbook(excel_buffer)
            sheet = pxl_doc['Blog import']
            images = sheet._images
            print(f"Nombre d'images trouvées : {len(images)}")
            i = 0
            translator = Translator(to_lang="fr")
            to_import = False
            for row in sheet.iter_rows():
                to_import = eval(row[23].value)
                if i > 0 and to_import:
                    blog, blog_fr = convert_row_to_blog_post(sheet,row, i, translator, self)
                    blog_post = self.env['blog.post'].with_context(lang='en_US').create(blog)
                    blog_post.with_context(lang='fr_FR').write({
                        'name': blog_fr['name'],
                        'subtitle': blog_fr['subtitle']
                    })
                    content = {
                        "en_US": blog['content'],
                        "fr_FR": blog_fr['content']
                    }
                    str_content = json.dumps(content)
                    req = """UPDATE blog_post p SET content = %s WHERE p.id = %s""" % ("'" + str_content.replace("'","''") + "'",blog_post.id)
                    self._cr.execute(req)
                i = i + 1
            return UserError(_('Fichier importé correctement'))
        except Exception as e:
            raise UserError(_('Erreur dans votre format de fichier') + str(e))